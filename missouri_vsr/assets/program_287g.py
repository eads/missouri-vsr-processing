from __future__ import annotations

import hashlib
import json
import re
from datetime import date, datetime, timezone
from pathlib import Path

import httpx
import pandas as pd
from dagster import AssetCheckResult, AssetKey, asset, asset_check
from openpyxl import load_workbook

from missouri_vsr.assets.reports import _build_agency_name_lookup
from missouri_vsr.cli.crosswalk import _normalize_name


SNAPSHOT_DIR = Path("data/src/287g")
SNAPSHOTS_LOG = SNAPSHOT_DIR / "snapshots.jsonl"
ICE_PAGE_URL = "https://www.ice.gov/identify-and-arrest/287g"

# appelson/Tracking_287g runs a daily scraper that mirrors the ICE xlsx into
# timestamped sheets_YYYYMMDD_HHMMSS/ folders. We use it as a reliable mirror
# because the ICE page itself is Akamai-protected and blocks plain HTTP.
APPELSON_REPO = "appelson/Tracking_287g"
APPELSON_SHEETS_API = f"https://api.github.com/repos/{APPELSON_REPO}/contents/sheets"
APPELSON_FOLDER_RE = re.compile(r"^sheets_\d{8}_\d{6}$")

# ICE filename: participatingAgenciesMMDDYYYY{am|pm}.xlsx
_FILENAME_RE = re.compile(
    r"participatingAgencies(?P<m>\d{2})(?P<d>\d{2})(?P<y>\d{4})(?P<period>am|pm)(?:_(?P<seq>\d+))?\.xlsx",
    re.IGNORECASE,
)


def _parse_snapshot_filename(name: str) -> tuple[date, str, int] | None:
    m = _FILENAME_RE.match(name)
    if not m:
        return None
    snap = date(int(m["y"]), int(m["m"]), int(m["d"]))
    seq = int(m["seq"] or 0)
    return snap, m["period"].lower(), seq


def _latest_snapshot_path(snapshot_dir: Path) -> Path:
    candidates = []
    for p in snapshot_dir.glob("participatingAgencies*.xlsx"):
        parsed = _parse_snapshot_filename(p.name)
        if parsed is None:
            continue
        snap_date, period, seq = parsed
        # pm sorts after am within a day
        candidates.append(((snap_date, 1 if period == "pm" else 0, seq), p))
    if not candidates:
        raise FileNotFoundError(
            f"No 287g snapshot found in {snapshot_dir}. "
            f"Manually download from {ICE_PAGE_URL} (see fetch_287g_snapshot_stub)."
        )
    candidates.sort(key=lambda item: item[0])
    return candidates[-1][1]


def _read_snapshot(xlsx_path: Path) -> pd.DataFrame:
    """Load the 287g xlsx and extract MOA hyperlink targets via openpyxl."""
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    expected = ["STATE", "LAW ENFORCEMENT AGENCY", "TYPE", "COUNTY", "SUPPORT TYPE", "SIGNED", "MOA"]
    for col in expected:
        if col not in headers:
            raise ValueError(f"287g snapshot missing expected column {col!r} (got: {headers})")
    moa_idx = headers.index("MOA")

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        values = {headers[i]: cell.value for i, cell in enumerate(row) if i < len(headers)}
        if not values.get("STATE"):
            continue
        moa_cell = row[moa_idx]
        moa_url = moa_cell.hyperlink.target if moa_cell.hyperlink else None
        moa_status = (str(moa_cell.value).strip() if moa_cell.value is not None else "") or None
        signed = values.get("SIGNED")
        if isinstance(signed, datetime):
            signed_str = signed.date().isoformat()
        elif isinstance(signed, date):
            signed_str = signed.isoformat()
        elif signed is None:
            signed_str = None
        else:
            signed_str = str(signed).strip() or None

        rows.append(
            {
                "state": str(values.get("STATE", "")).strip(),
                "ice_agency_name": str(values.get("LAW ENFORCEMENT AGENCY", "")).strip(),
                "ice_agency_type": str(values.get("TYPE", "")).strip() or None,
                "county": str(values.get("COUNTY", "")).strip() or None,
                "support_type": str(values.get("SUPPORT TYPE", "")).strip() or None,
                "signed_date": signed_str,
                "moa_status": moa_status,
                "moa_url": moa_url,
            }
        )
    return pd.DataFrame(rows)


def _resolve_canonical(ice_name: str, lookup: dict[str, str]) -> str | None:
    return lookup.get(_normalize_name(ice_name))


def _build_canonical_self_lookup(canonical_parquet: Path) -> dict[str, str]:
    """{normalized_canonical_name: canonical_name} from the canonical_combined parquet.

    Lets ICE names match canonical names directly when normalization alone is enough,
    without needing a manual crosswalk row for every agency.
    """
    if not canonical_parquet.exists():
        return {}
    df = pd.read_parquet(canonical_parquet, columns=["agency"])
    out: dict[str, str] = {}
    for name in df["agency"].dropna().astype(str).unique():
        out.setdefault(_normalize_name(name), name)
    return out


def _list_appelson_sheets_folders(client: httpx.Client) -> list[str]:
    """Return appelson sheets_* folder names, newest last (lexical sort works for the timestamp format)."""
    r = client.get(APPELSON_SHEETS_API, headers={"Accept": "application/vnd.github+json"})
    r.raise_for_status()
    names = [
        item["name"]
        for item in r.json()
        if item.get("type") == "dir" and APPELSON_FOLDER_RE.match(item.get("name", ""))
    ]
    names.sort()
    return names


def _find_xlsx_in_folder(client: httpx.Client, folder: str) -> dict:
    """Return GitHub contents entry for the xlsx file inside an appelson sheets_* folder."""
    r = client.get(f"{APPELSON_SHEETS_API}/{folder}", headers={"Accept": "application/vnd.github+json"})
    r.raise_for_status()
    for item in r.json():
        name = item.get("name", "")
        if item.get("type") == "file" and _parse_snapshot_filename(name) is not None:
            return item
    raise FileNotFoundError(f"No participatingAgencies*.xlsx file found in appelson folder {folder}")


def _fetch_latest_appelson_snapshot(snapshot_dir: Path, log) -> dict:
    """Pull the most recent appelson mirror snapshot into snapshot_dir.

    Idempotent on filename: if the target file already exists locally, skip download
    and skip appending to snapshots.jsonl. Returns a metadata dict describing the
    chosen snapshot (whether or not it was just downloaded).
    """
    snapshot_dir.mkdir(parents=True, exist_ok=True)
    with httpx.Client(timeout=30.0, follow_redirects=True) as client:
        folders = _list_appelson_sheets_folders(client)
        if not folders:
            raise RuntimeError(f"No sheets_* folders found in {APPELSON_REPO}")
        latest_folder = folders[-1]
        log.info("appelson latest folder: %s", latest_folder)

        entry = _find_xlsx_in_folder(client, latest_folder)
        filename = entry["name"]
        target = snapshot_dir / filename
        download_url = entry["download_url"]

        if target.exists():
            log.info("Snapshot %s already present locally — skipping download", filename)
            return {
                "filename": filename,
                "appelson_folder": latest_folder,
                "download_url": download_url,
                "downloaded": False,
                "path": target,
            }

        log.info("Downloading %s from %s", filename, download_url)
        r = client.get(download_url)
        r.raise_for_status()
        content = r.content
        target.write_bytes(content)
        sha256 = hashlib.sha256(content).hexdigest()

        parsed = _parse_snapshot_filename(filename)
        snap_date, period, _ = parsed  # safe: _find_xlsx_in_folder filtered on parseability
        record = {
            "filename": filename,
            "fetched_at": datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds"),
            "source_url_observed": download_url,
            "snapshot_date": snap_date.isoformat(),
            "snapshot_period": period,
            "sha256": sha256,
            "fetched_by": "appelson_mirror",
            "notes": f"Mirrored from {APPELSON_REPO} folder {latest_folder}.",
        }
        log_path = snapshot_dir / "snapshots.jsonl"
        with log_path.open("a") as f:
            f.write(json.dumps(record) + "\n")

        return {
            "filename": filename,
            "appelson_folder": latest_folder,
            "download_url": download_url,
            "downloaded": True,
            "sha256": sha256,
            "path": target,
        }


@asset(
    name="fetch_287g_snapshot",
    group_name="program_287g",
    required_resource_keys={"data_dir_source"},
    description=(
        "Pulls the most recent 287g participating-agencies xlsx from the "
        f"{APPELSON_REPO} daily-scraper mirror into data/src/287g/. "
        "Idempotent on filename: re-running when the latest mirror snapshot is "
        "already local is a no-op. Always materialize this before ice_287g_latest."
    ),
)
def fetch_287g_snapshot(context) -> str:
    src_dir = Path(context.resources.data_dir_source.get_path())
    snapshot_dir = src_dir / "287g"
    result = _fetch_latest_appelson_snapshot(snapshot_dir, context.log)
    context.add_output_metadata(
        {
            "filename": result["filename"],
            "appelson_folder": result["appelson_folder"],
            "downloaded": result["downloaded"],
            "source_url": result["download_url"],
        }
    )
    return result["filename"]


@asset(
    name="ice_287g_latest",
    group_name="program_287g",
    deps=[AssetKey("canonical_combined_parquet"), AssetKey("fetch_287g_snapshot")],
    required_resource_keys={"data_dir_source", "data_dir_processed"},
    description=(
        "Missouri rows from the latest ICE 287g participating-agencies snapshot, "
        "joined to canonical agency names via agency_crosswalk.csv + the canonical "
        "agency list from canonical_combined.parquet."
    ),
)
def ice_287g_latest(context) -> pd.DataFrame:
    src_dir = Path(context.resources.data_dir_source.get_path())
    snapshot_dir = src_dir / "287g"
    xlsx_path = _latest_snapshot_path(snapshot_dir)
    parsed = _parse_snapshot_filename(xlsx_path.name)
    snap_date, period, seq = parsed  # safe: _latest_snapshot_path filtered for parseability
    context.log.info("Reading 287g snapshot %s (date=%s period=%s)", xlsx_path.name, snap_date, period)

    df = _read_snapshot(xlsx_path)
    mo = df[df["state"].str.upper() == "MISSOURI"].copy()
    context.log.info("287g rows: %d total, %d Missouri", len(df), len(mo))

    crosswalk_path = src_dir / "agency_crosswalk.csv"
    crosswalk_lookup = _build_agency_name_lookup(crosswalk_path)
    processed_dir = Path(context.resources.data_dir_processed.get_path())
    self_lookup = _build_canonical_self_lookup(processed_dir / "canonical_combined.parquet")
    # Crosswalk wins on normalization key collisions (it's the curated mapping).
    combined_lookup = {**self_lookup, **crosswalk_lookup}
    mo["agency_canonical"] = mo["ice_agency_name"].apply(
        lambda n: _resolve_canonical(n, combined_lookup)
    )

    unresolved = mo[mo["agency_canonical"].isna()]["ice_agency_name"].unique().tolist()
    if unresolved:
        context.log.warning(
            "287g: %d ICE agency names did not resolve to canonical names: %s",
            len(unresolved), unresolved,
        )

    mo["snapshot_date"] = snap_date.isoformat()
    mo["snapshot_period"] = period
    mo["snapshot_filename"] = xlsx_path.name

    out_dir = Path(context.resources.data_dir_processed.get_path())
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "ice_287g_latest.parquet"
    mo.to_parquet(out_path, index=False)
    context.log.info("Wrote %s (%d rows)", out_path, len(mo))
    context.add_output_metadata(
        {
            "snapshot_filename": xlsx_path.name,
            "snapshot_date": snap_date.isoformat(),
            "mo_rows": len(mo),
            "distinct_mo_agencies": int(mo["ice_agency_name"].nunique()),
            "unresolved_count": len(unresolved),
        }
    )
    return mo


@asset_check(asset=ice_287g_latest, name="all_mo_agencies_resolved")
def check_all_mo_agencies_resolved(ice_287g_latest: pd.DataFrame) -> AssetCheckResult:
    """Every Missouri 287g agency should map to a canonical name via the crosswalk.

    A miss means the ICE name spelling is new — add a row to agency_crosswalk.csv.
    """
    if ice_287g_latest.empty:
        return AssetCheckResult(passed=True, metadata={"note": "empty snapshot"})
    unresolved = (
        ice_287g_latest[ice_287g_latest["agency_canonical"].isna()]["ice_agency_name"]
        .unique()
        .tolist()
    )
    return AssetCheckResult(
        passed=not unresolved,
        metadata={"unresolved_ice_names": unresolved, "unresolved_count": len(unresolved)},
    )


